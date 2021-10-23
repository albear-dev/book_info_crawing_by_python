from openpyxl import Workbook
from page.page_info import PageInfo
import os
from queue import Queue
import threading
from crlogging.cr_logger import CRLogger
from openpyxl.styles import PatternFill
from tkinter import messagebox

class ExcelHandler:
    __logger = CRLogger.get_logger(__name__)

    def __init__(self, ui_option):
        self.page_processing_queue = Queue()
        self.write_wb = Workbook()
        self.write_wb.remove_sheet(self.write_wb.get_sheet_by_name('Sheet'))
        self.path = ui_option.file_path
        self.filename = ui_option.file_name
        self.highlight_keyword = ui_option.highlight_keyword
        self.worker = threading.Thread(target=self.excel_write)
        self.worker.start()
        #self.worker.join()

    def add_page(self, page: PageInfo):
        self.page_processing_queue.put(page)

    def excel_write(self):
        while True:
            page = self.page_processing_queue.get()
            self.__logger.info("get page object [%s]", page)

            if not page:
                break

            self.__logger.info("Start excel write page [%s]", page.title)

            # create workbook
            write_ws = self.write_wb.create_sheet(page.title)
            self.__logger.info("Sheet created [%s]", page.title)

            # write title
            write_ws.append(["순위", "상품번호", "상품명", "정가", "할인가", "출판사", "저자", "출간일", "판매지수"])
            for data in page.book_info_collection:
                # write data
                write_ws.append([
                    data.rank,
                    data.prod_no,
                    data.prod_nm,
                    data.price,
                    data.sale_price,
                    data.publisher,
                    data.author,
                    data.release_date,
                    data.selling_score
                ])

                max_col = 9
                if self.highlight_keyword:
                    keywords = self.highlight_keyword.split(",")
                    for keyword in keywords:
                        if data.prod_nm and data.prod_nm.find(keyword) > -1:
                            for rows in write_ws.iter_rows(min_row=write_ws.max_row, max_row=write_ws.max_row, min_col=1, max_col=max_col):
                                for cell in rows:
                                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                            write_ws.cell(row=write_ws.max_row, column=3).fill = PatternFill(start_color='FFBB33', end_color='FFBB33', fill_type='solid')
                        if data.publisher and data.publisher.find(keyword) > -1:
                            for rows in write_ws.iter_rows(min_row=write_ws.max_row, max_row=write_ws.max_row, min_col=1, max_col=max_col):
                                for cell in rows:
                                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                            write_ws.cell(row=write_ws.max_row, column=6).fill = PatternFill(start_color='FFBB33', end_color='FFBB33', fill_type='solid')
                        if data.author and data.author.find(keyword) > -1:
                            for rows in write_ws.iter_rows(min_row=write_ws.max_row, max_row=write_ws.max_row, min_col=1, max_col=max_col):
                                for cell in rows:
                                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                            write_ws.cell(row=write_ws.max_row, column=7).fill = PatternFill(start_color='FFBB33', end_color='FFBB33', fill_type='solid')

            self.__logger.info("Finish write data to sheet [%s] count[%d]", page.title, len(page.book_info_collection))

        save_path = os.path.join(self.path, self.filename)
        self.write_wb.save(save_path)
        self.__logger.info("Excel file save done. [%s]", save_path)
        messagebox.showinfo("안내", "파일 생성이 완료 되었습니다."+ save_path)

